"""
Authoritative validation of an uploaded procurement file.

Runs in the worker against the downloaded file (the browser's pre-check is only
for UX). Column detection uses the SAME synonym mapping as the pipeline ingest,
so any file the pipeline can actually process passes the gate — and anything it
can't is rejected with a specific, human-readable message that points the
analyst at the required format. Never surfaces a stack trace.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from django.conf import settings

from pipeline.ingest import detect_columns


@dataclass
class ValidationResult:
    ok: bool
    error_code: str = ""
    message: str = ""


def _ext(filename: str) -> str:
    return Path(filename).suffix.lower()


def _human_mb(num_bytes: int) -> str:
    return f"{num_bytes / (1024 * 1024):.1f} MB"


def _xlsx_bomb_check(path: Path) -> bool:
    """True if the .xlsx decompresses to within the configured limit."""
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            total = sum(info.file_size for info in zf.infolist())
    except zipfile.BadZipFile:
        return True  # not a valid zip; let the normal reader surface 'unreadable'
    return total <= settings.MAX_XLSX_UNCOMPRESSED_BYTES


def _header_and_has_rows(path: Path, ext: str) -> tuple[list[str], bool]:
    """Return (header, has_at_least_one_data_row)."""
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            header = next(reader, [])
            has_row = any(any((cell or "").strip() for cell in row) for row in reader)
            return header, has_row
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = [str(c) for c in next(rows, ()) if c is not None]
            has_row = any(any(c is not None and str(c).strip() for c in row) for row in rows)
            return header, has_row
        finally:
            wb.close()
    return [], False


def validate_upload(path: Path, filename: str, size: int) -> ValidationResult:
    ext = _ext(filename)

    if ext not in settings.ALLOWED_UPLOAD_EXTENSIONS:
        allowed = " or ".join(settings.ALLOWED_UPLOAD_EXTENSIONS)
        return ValidationResult(
            False,
            "bad_extension",
            f"Unsupported file type '{ext or 'unknown'}'. Please upload a {allowed} file.",
        )

    if size > settings.MAX_UPLOAD_BYTES:
        limit = _human_mb(settings.MAX_UPLOAD_BYTES)
        return ValidationResult(
            False,
            "too_large",
            f"File is {_human_mb(size)}, which exceeds the {limit} limit. "
            "Please split or compress it.",
        )

    if ext == ".xlsx" and not _xlsx_bomb_check(path):
        return ValidationResult(
            False,
            "xlsx_too_large_uncompressed",
            "This Excel file expands to an unexpectedly large size and was rejected "
            "for safety. Please export to CSV or a smaller workbook.",
        )

    try:
        header, has_rows = _header_and_has_rows(path, ext)
    except Exception:
        return ValidationResult(
            False,
            "unreadable",
            "We couldn't read this file. Please make sure it is a valid, "
            "uncorrupted CSV or Excel file with column headers in the first row.",
        )

    if not header:
        return ValidationResult(
            False,
            "no_header",
            "This file has no column headers in the first row. Add a header row "
            "(e.g. product, spend, quantity, unit) and re-upload.",
        )

    mapping = detect_columns(header)

    if "product" not in mapping:
        return ValidationResult(
            False,
            "missing_product",
            "We couldn't find a product/description column. Add a column named "
            "'product' (or 'item' / 'description') naming each food purchased.",
        )

    if "spend" not in mapping and "quantity" not in mapping:
        return ValidationResult(
            False,
            "missing_value_column",
            "We need either a 'spend' (dollar amount) column or a 'quantity' "
            "column to estimate impact. Please add at least one and re-upload.",
        )

    if not has_rows:
        return ValidationResult(
            False,
            "empty",
            "This file has a header row but no data rows. Please add your "
            "purchasing line items and re-upload.",
        )

    return ValidationResult(True)
